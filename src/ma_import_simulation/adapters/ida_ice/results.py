"""Kleine, nachvollziehbare Leser fuer ausgewaehlte IDA-ICE-Ergebnisartefakte.

Die Funktionen berechnen keine Kennwerte. Sie lesen nur explizit genannte
Ergebnisdateien und halten Herkunft und Pruefstatus fest.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

ADAPTER_KEY = "ida_ice"
SCHEMA_VERSION = "1.0"
ALLOWED_ARTIFACT_SUFFIXES = {".prn": "prn", ".html": "html", ".htm": "html", ".xlsx": "excel"}
PROTECTED_SUFFIXES = {".idm", ".idc"}
KNOWN_COHORTS = {"5Z", "29Z", "ALT"}
DEFAULT_MANIFEST_NAME = "ida_result_manifest.json"


@dataclass(frozen=True, slots=True)
class IdaDiagnostic:
    """Hinweis, warum eine Quelle angenommen oder abgelehnt wurde."""

    code: str
    message: str
    path: Path | None = None


@dataclass(frozen=True, slots=True)
class IdaArtifact:
    """Positiv erkannte, erlaubte Ergebnisdatei."""

    path: Path
    kind: str
    sha256: str
    status: str = "accepted"


@dataclass(frozen=True, slots=True)
class RawIdaResult:
    """Rohwert mit den fuer eine spaetere Zuordnung benoetigten Metadaten."""

    schema_version: str
    run_id: str
    variant_id: str
    model_id: str
    zone_id: str | None
    scope: str
    metric: str
    unit: str | None
    sign: str | None
    time: float | None
    value: float | None
    provenance: str
    source_hash: str
    status: str


@dataclass(frozen=True, slots=True)
class StandardizedIdaResult:
    """Nicht aggregierter, einheitlich beschriebener Ergebniswert."""

    schema_version: str
    run_id: str
    variant_id: str
    model_id: str
    zone_id: str | None
    scope: str
    metric: str
    unit: str | None
    sign: str | None
    time: float | None
    value: float | None
    provenance: str
    source_hash: str
    status: str

    @classmethod
    def from_raw(cls, raw: RawIdaResult) -> "StandardizedIdaResult":
        """Uebernimmt einen Rohwert unveraendert in den Standardvertrag."""
        return cls(**{name: getattr(raw, name) for name in cls.__dataclass_fields__})


@dataclass(frozen=True, slots=True)
class IdaPackage:
    """Explizit per Manifest beschriebene Ergebniskohorte."""

    root: Path
    cohort: str
    artifacts: tuple[IdaArtifact, ...]
    diagnostics: tuple[IdaDiagnostic, ...] = field(default_factory=tuple)


def sha256_file(path: str | Path) -> str:
    """Ermittelt die SHA-256-Pruefsumme einer einzelnen Datei."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def detect_ida_artifact(path: str | Path) -> tuple[IdaArtifact | None, tuple[IdaDiagnostic, ...]]:
    """Erkennt nur erlaubte Ergebnisdateien; IDA-Modellquellen bleiben ungelesen."""
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix in PROTECTED_SUFFIXES:
        return None, (IdaDiagnostic("protected_source", "IDM/IDC-Inhalte werden nicht verarbeitet.", source),)
    if not source.is_file():
        return None, (IdaDiagnostic("not_a_file", "Die Quelle ist keine vorhandene Datei.", source),)
    kind = ALLOWED_ARTIFACT_SUFFIXES.get(suffix)
    if kind is None:
        return None, (IdaDiagnostic("unsupported_artifact", "Die Datei ist kein erlaubtes IDA-Ergebnisartefakt.", source),)
    return IdaArtifact(path=source, kind=kind, sha256=sha256_file(source)), ()


def inspect_ida_package(root: str | Path, manifest_name: str = DEFAULT_MANIFEST_NAME) -> IdaPackage:
    """Liest ausschließlich das Manifest direkt unter ``root`` und seine Eintraege.

    Das Manifest ist JSON mit ``cohort`` (``5Z``, ``29Z`` oder ``ALT``) und
    ``artifacts`` als Liste relativer Dateinamen. Es findet kein Dateiscan statt.
    """
    package_root = Path(root).resolve()
    manifest_path = package_root / manifest_name
    if not package_root.is_dir() or not manifest_path.is_file():
        raise ValueError("IDA-Paket braucht eine explizite Wurzel mit Manifestdatei.")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError("IDA-Paketmanifest ist kein gueltiges JSON.") from exc
    cohort = manifest.get("cohort")
    entries = manifest.get("artifacts")
    if cohort not in KNOWN_COHORTS:
        raise ValueError("IDA-Paketmanifest braucht genau eine bekannte Kohorte: 5Z, 29Z oder ALT.")
    if not isinstance(entries, list) or not entries:
        raise ValueError("IDA-Paketmanifest braucht mindestens ein explizites Artefakt.")

    artifacts: list[IdaArtifact] = []
    diagnostics: list[IdaDiagnostic] = []
    seen: set[Path] = set()
    for entry in entries:
        if not isinstance(entry, str):
            diagnostics.append(IdaDiagnostic("invalid_manifest_entry", "Artefaktname im Manifest ist kein Text."))
            continue
        candidate = (package_root / entry).resolve()
        if package_root not in candidate.parents or candidate in seen:
            diagnostics.append(IdaDiagnostic("unauthorized_source", "Manifest verweist ausserhalb der Paketwurzel oder doppelt.", candidate))
            continue
        seen.add(candidate)
        artifact, item_diagnostics = detect_ida_artifact(candidate)
        diagnostics.extend(item_diagnostics)
        if artifact is not None:
            artifacts.append(artifact)
    if not artifacts:
        raise ValueError("IDA-Paket enthält keine erlaubten, explizit referenzierten Ergebnisartefakte.")
    return IdaPackage(package_root, cohort, tuple(artifacts), tuple(diagnostics))


def parse_prn_file(path: str | Path) -> tuple[tuple[str, ...], tuple[tuple[float, ...], ...]]:
    """Liest eine PRN-Tabelle mit ``#``-Kopfzeile und variabler Zeitspalte.

    Die erste numerische Spalte wird als Zeit beibehalten; es wird bewusst keine
    stündliche Umrechnung oder Kennwertbildung vorgenommen.
    """
    artifact, diagnostics = detect_ida_artifact(path)
    if artifact is None or artifact.kind != "prn":
        raise ValueError(diagnostics[0].message if diagnostics else "Erwartet wird eine PRN-Datei.")
    lines = artifact.path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    header_line = next((line for line in lines if line.lstrip().startswith("#")), None)
    if header_line is None:
        raise ValueError("PRN-Datei braucht eine Kopfzeile mit '#'.")
    header = tuple(header_line.lstrip()[1:].strip().split())
    if not header:
        raise ValueError("PRN-Kopfzeile enthält keine Spaltennamen.")
    rows: list[tuple[float, ...]] = []
    for line in lines[lines.index(header_line) + 1 :]:
        parts = line.split()
        if not parts or line.lstrip().startswith("#"):
            continue
        try:
            values = tuple(float(value.replace(",", ".")) for value in parts)
        except ValueError as exc:
            raise ValueError("PRN-Datenzeile enthaelt nichtnumerische Werte.") from exc
        if len(values) != len(header):
            raise ValueError("PRN-Datenzeile hat nicht die Spaltenzahl der Kopfzeile.")
        rows.append(values)
    if not rows:
        raise ValueError("PRN-Datei enthält keine numerischen Datenzeilen.")
    return header, tuple(rows)


class _ReportHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.meta: dict[str, str] = {}
        self.tables: list[tuple[tuple[str, ...], ...]] = []
        self._stack: list[dict[str, Any]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "meta" and attributes.get("name") and attributes.get("content"):
            self.meta[attributes["name"]] = attributes["content"]
        elif tag == "table":
            self._stack.append({"rows": [], "row": None, "cell": None})
        elif tag == "tr" and self._stack:
            self._finish_row()
            self._stack[-1]["row"] = []
        elif tag in {"th", "td"} and self._stack and self._stack[-1]["row"] is not None:
            self._finish_cell()
            self._stack[-1]["cell"] = []

    def handle_data(self, data: str) -> None:
        if self._stack and self._stack[-1]["cell"] is not None:
            self._stack[-1]["cell"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self._stack:
            self._finish_cell()
        elif tag == "tr" and self._stack:
            self._finish_row()
        elif tag == "table" and self._stack:
            self._finish_row()
            context = self._stack.pop()
            if context["rows"]:
                self.tables.append(tuple(context["rows"]))

    def _finish_cell(self) -> None:
        context = self._stack[-1]
        if context["cell"] is not None and context["row"] is not None:
            context["row"].append(" ".join(context["cell"]).strip())
            context["cell"] = None

    def _finish_row(self) -> None:
        context = self._stack[-1]
        self._finish_cell()
        if context["row"]:
            context["rows"].append(tuple(context["row"]))
        context["row"] = None


def parse_html_report(path: str | Path) -> tuple[dict[str, str], tuple[tuple[tuple[str, ...], ...], ...]]:
    """Extrahiert Meta-Tags und tatsächlich vorhandene HTML-Tabellen ohne Deutung."""
    artifact, diagnostics = detect_ida_artifact(path)
    if artifact is None or artifact.kind != "html":
        raise ValueError(diagnostics[0].message if diagnostics else "Erwartet wird ein HTML-Bericht.")
    parser = _ReportHtmlParser()
    parser.feed(artifact.path.read_text(encoding="utf-8-sig", errors="replace"))
    return parser.meta, tuple(parser.tables)


def read_excel_metadata(path: str | Path) -> dict[str, Any]:
    """Liest Arbeitsmappen-Metadaten einer explizit ausgewählten XLSX-Datei."""
    artifact, diagnostics = detect_ida_artifact(path)
    if artifact is None or artifact.kind != "excel":
        raise ValueError(diagnostics[0].message if diagnostics else "Erwartet wird eine XLSX-Datei.")
    from openpyxl import load_workbook

    workbook = load_workbook(artifact.path, read_only=True, data_only=True)
    try:
        properties = workbook.properties
        return {
            "sheet_names": tuple(workbook.sheetnames),
            "title": properties.title,
            "creator": properties.creator,
            "created": properties.created.isoformat() if properties.created else None,
            "modified": properties.modified.isoformat() if properties.modified else None,
            "source_hash": artifact.sha256,
        }
    finally:
        workbook.close()
