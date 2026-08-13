"""Leser fuer explizit ausgewaehlte IDA-Zonenmetadaten-Arbeitsmappen."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class ZoneMetadata:
    zone: str
    group: str
    area_m2: float | None
    multiplier: float | None
    max_supply_airflow_l_s_m2: float | None
    max_exhaust_airflow_l_s_m2: float | None


def read_zone_metadata(path: str | Path) -> tuple[ZoneMetadata, ...]:
    """Liest die belegten Abschnitte `Zonen`, `Zonen - total` und Sollwerte."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()

    zones = _section_rows(rows, "zonen")
    totals = {str(row.get("name")): row for row in _section_rows(rows, "zonen_total")}
    setpoints = {str(row.get("name")): row for row in _section_rows(rows, "zonen_sollwerte")}
    result: list[ZoneMetadata] = []
    for row in zones:
        name = str(row.get("name") or "").strip()
        if not name or name.casefold().startswith(("gesamt", "total")):
            continue
        total = totals.get(name, {})
        setpoint = setpoints.get(name, {})
        result.append(
            ZoneMetadata(
                zone=name,
                group=str(row.get("gruppe") or "").strip(),
                area_m2=_number(row.get("bodenflache_m2")),
                multiplier=_number(total.get("zonenmultiplikator")),
                max_supply_airflow_l_s_m2=_number(setpoint.get("max_vvs_zuluft_l_s_m2")),
                max_exhaust_airflow_l_s_m2=_number(setpoint.get("max_vvs_abluft_l_s_m2")),
            )
        )
    return tuple(result)


def _section_rows(rows: list[tuple], section_key: str) -> list[dict[str, object]]:
    target = _key(section_key)
    for index, row in enumerate(rows):
        first = next((value for value in row if value is not None), None)
        if first is None or _key(str(first)) != target:
            continue
        if index + 1 >= len(rows):
            return []
        headers = tuple(_key(str(value)) if value is not None else "" for value in rows[index + 1])
        data: list[dict[str, object]] = []
        for values in rows[index + 2 :]:
            if not any(value is not None for value in values):
                break
            first_value = next((value for value in values if value is not None), "")
            if isinstance(first_value, str) and _key(first_value) in {
                "zonen_total",
                "zonen_sollwerte",
                "used_template",
                "lokale_heiz_kuhlelemente",
                "terminals",
                "interne_warmequellen",
                "zeitplane",
                "interne_massen",
            }:
                break
            data.append({header: value for header, value in zip(headers, values, strict=False) if header})
        return data
    return []


def _key(value: str) -> str:
    normalized = value.casefold().replace("²", "2").replace("ä", "a").replace("ö", "o").replace("ü", "u")
    normalized = normalized.replace("ß", "ss").replace("\xad", "")
    return re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
