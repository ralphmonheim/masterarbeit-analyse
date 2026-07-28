"""Lesender Zugriff auf die lokalen Excel-Kataloge fuer Gebaeudeinhalte."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_BUILDING_EXCEL_CATALOGS = {
    "components": Path("data/catalogs/components/demo_masterarbeit_bauteilkatalog.xlsx"),
    "materials": Path("data/catalogs/materials/demo_masterarbeit_materialkatalog.xlsx"),
    "products": Path("data/catalogs/products/demo_masterarbeit_produktkatalog.xlsx"),
}
_ID_PREFIXES = {
    "components": ("BP-", "DP-", "AW-", "IW-", "DA-"),
    "materials": ("MAT-",),
    "products": ("PROD-",),
}
_ID_HEADERS = {
    "components": "Bauteil-ID",
    "materials": "Material-ID",
    "products": "Produkt-ID",
}


@dataclass(frozen=True, slots=True)
class BuildingExcelCatalog:
    """Kleine, unveraenderte Sicht auf die Uebersicht einer Arbeitsmappe."""

    catalog_type: str
    source_path: Path
    source_sha256: str
    rows: tuple[dict[str, object], ...]


def load_building_excel_catalog(
    catalog_type: str,
    source_path: str | Path | None = None,
) -> BuildingExcelCatalog:
    """Laedt ausschliesslich Datensaetze aus dem Blatt ``Uebersicht``."""
    if catalog_type not in DEFAULT_BUILDING_EXCEL_CATALOGS:
        raise ValueError(f"Unbekannter Excel-Katalogtyp: {catalog_type}")
    path = Path(source_path or DEFAULT_BUILDING_EXCEL_CATALOGS[catalog_type])
    if not path.is_file():
        raise FileNotFoundError(f"Excel-Katalog nicht gefunden: {path}")
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Übersicht" not in book.sheetnames:
            raise ValueError(f"Excel-Katalog braucht das Blatt 'Übersicht': {path}")
        sheet = book["Übersicht"]
        headers = tuple(
            str(value).strip() if value is not None else ""
            for value in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
        )
        if not headers or not headers[0]:
            raise ValueError(f"Excel-Katalog hat keine Tabellenkoepfe: {path}")
        nonempty_headers = tuple(header for header in headers if header)
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise ValueError(f"Excel-Katalog enthaelt doppelte Tabellenkoepfe: {path}")
        expected_id_header = _ID_HEADERS[catalog_type]
        if headers[0] != expected_id_header:
            raise ValueError(
                f"Excel-Katalog braucht '{expected_id_header}' als erste Pflichtspalte: {path}"
            )
        prefixes = _ID_PREFIXES[catalog_type]
        rows = tuple(
            {
                header: value
                for header, value in zip(headers, raw_row, strict=True)
                if header
            }
            for raw_row in sheet.iter_rows(min_row=5, values_only=True)
            if isinstance(raw_row[0], str) and raw_row[0].startswith(prefixes)
        )
        record_ids = [str(row[expected_id_header]) for row in rows]
        if len(record_ids) != len(set(record_ids)):
            raise ValueError(f"Excel-Katalog enthaelt doppelte Datensatz-IDs: {path}")
    finally:
        book.close()
    return BuildingExcelCatalog(
        catalog_type=catalog_type,
        source_path=path.resolve(),
        source_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        rows=rows,
    )
