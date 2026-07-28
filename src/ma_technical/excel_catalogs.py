"""Lesender V1-Adapter fuer lokale Techniksystem-Excel-Kataloge."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class TechnicalExcelCatalog:
    source_path: Path
    source_sha256: str
    id_column: str
    rows: tuple[dict[str, object], ...]


def load_technical_excel_catalog(source_path: str | Path) -> TechnicalExcelCatalog:
    """Liest eine Uebersichtstabelle, ohne Produktdaten zu erfinden."""
    path = Path(source_path)
    if not path.is_file():
        raise FileNotFoundError(f"Techniksystem-Excel-Katalog fehlt: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Übersicht" not in workbook.sheetnames:
            raise ValueError("Techniksystem-Katalog braucht das Blatt 'Übersicht'.")
        sheet = workbook["Übersicht"]
        header_row, headers = _find_headers(sheet)
        nonempty_headers = tuple(header for header in headers if header)
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise ValueError("Techniksystem-Katalog enthaelt doppelte Ueberschriften.")
        id_column = next(
            (
                header
                for header in headers
                if header.casefold() in {"system-id", "techniksystem-id", "paket-id"}
            ),
            "",
        )
        if not id_column:
            raise ValueError(
                "Techniksystem-Katalog braucht eine Spalte System-ID, Techniksystem-ID oder Paket-ID."
            )
        rows = tuple(
            {
                header: value
                for header, value in zip(headers, raw_row, strict=True)
                if header
            }
            for raw_row in sheet.iter_rows(min_row=header_row + 1, values_only=True)
            if raw_row[headers.index(id_column)] not in {None, ""}
        )
        record_ids = [str(row[id_column]) for row in rows]
        if len(record_ids) != len(set(record_ids)):
            raise ValueError("Techniksystem-Katalog enthaelt doppelte Datensatz-IDs.")
    finally:
        workbook.close()
    return TechnicalExcelCatalog(
        source_path=path.resolve(),
        source_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        id_column=id_column,
        rows=rows,
    )


def technical_catalog_record_status(row: dict[str, object]) -> tuple[bool, bool, str]:
    """Liefert aktiv, validiert und einen sichtbaren Status."""
    normalized = {str(key).casefold(): value for key, value in row.items()}
    active_value = normalized.get("aktiv", True)
    active = active_value not in {False, "0", "nein", "Nein", "inactive", "inaktiv"}
    status_value = next(
        (
            value
            for key, value in normalized.items()
            if key in {"validierungsstatus", "prüfstatus", "pruefstatus", "status"}
        ),
        "",
    )
    status = str(status_value).strip()
    validated = status.casefold() in {
        "validated",
        "validiert",
        "freigegeben",
        "released",
        "active",
        "aktiv",
    }
    return active, validated, status or "nicht validiert"


def _find_headers(sheet) -> tuple[int, tuple[str, ...]]:
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        values = next(
            sheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                values_only=True,
            )
        )
        headers = tuple(
            str(value).strip() if value is not None else ""
            for value in values
        )
        if any(
            header.casefold() in {"system-id", "techniksystem-id", "paket-id"}
            for header in headers
        ):
            return row_number, headers
    raise ValueError("Techniksystem-Katalog enthaelt keine erkennbare Ueberschrift.")
