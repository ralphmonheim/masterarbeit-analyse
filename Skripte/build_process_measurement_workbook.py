"""Ergaenzt die Prozessmessungs-Arbeitsmappe um die vereinbarten Auswertungen.

Die Datei liegt bewusst in der Arbeitsablage und nicht im Repository. Dieses
Skript veraendert keine bestehenden Blaetter, sondern fuegt nur neue Vorlagen
hinzu bzw. aktualisiert sie bei einem erneuten Lauf.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = Path(
    r"C:\Users\ralph\Documents\Master\5.Semester\Masterarbeit - lokal"
    r"\TEIL1_Fach-Anwendungskompetenz\260524_Masterarbeit_Arbeitsablage"
    r"\04_Teil2_Prozessinnovation\Prozessmessung"
    r"\Prozesskostenvergleich_Manuell_vs_Automatisiert.xlsx"
)

SHEETS = [
    "00_Gesamtübersicht",
    "01_PreProcess_Detail",
    "02_PreProcess_Übersicht",
    "03_Kernprozess_Übersicht",
    "04_Kernprozess_Einzelwerte",
    "05_PostProcess_Vorlage",
    "06_Kosten",
    "08_Messannahmen",
    "09_Quellenregister",
]

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
OPEN_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="B7B7B7")


def reset_sheet(workbook, name: str):
    """Erzeugt ein Auswertungsblatt neu; historische Ausgangsblätter bleiben erhalten."""
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    return workbook.create_sheet(name)


def title(sheet, text: str, subtitle: str, columns: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = sheet.cell(1, 1, text)
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    sheet.cell(2, 1, subtitle).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[2].height = 32
    sheet.freeze_panes = "A5"


def header(sheet, row: int, values: list[str]) -> None:
    for column, value in enumerate(values, 1):
        cell = sheet.cell(row, column, value)
        cell.font = Font(bold=True)
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[row].height = 30


def table_style(sheet, start_row: int, end_row: int, columns: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=columns):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)


def widths(sheet, values: list[int]) -> None:
    for column, width in enumerate(values, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def preprocess_detail(workbook) -> None:
    ws = reset_sheet(workbook, "01_PreProcess_Detail")
    title(ws, "PreProcess – Detailansicht", "Einzelne Tätigkeiten für beide Prozessmodi. Gelbe Felder sind manuell zu ergänzen; orange Felder sind bewusst offen.", 9)
    header(ws, 4, ["Modus", "ID", "Schritt", "Zeitart", "Dauer (min)", "Herkunft", "Status", "Quelle / Register-ID", "Nachweis / Erläuterung", "Variantenbezug"])
    rows = [
        ["Manuell", "M-P01", "Eingabemodule vollständig anlegen", "aktive Arbeitszeit", 150, "geschätzt", "vorläufig", "Q-003", "1–4 h für eine neue Variante; Mittelwert nur als Testwert", "neue Variante"],
        ["Manuell", "M-P02", "Parameterbereiche und Variantengenerierung", "aktive Arbeitszeit", None, "offen", "offen", "Q-004", "Noch nicht separat erhoben; nicht zusätzlich zum Gesamtwert zählen.", "neue Variante"],
        ["Manuell", "M-P03", "Prüfung der Eingaben", "Prüfzeit", None, "offen", "offen", "Q-004", "Noch nicht separat erhoben.", "neue Variante"],
        ["Automatisiert", "A-P01", "Projekt-Workspace anlegen", "Maschinenzeit", 0.0004023, "beobachtet", "gemessen", "Q-005", "BENCH-C278F4CF; 0,024139 s", "Benchmark"],
        ["Automatisiert", "A-P02", "PreProcess ausführen", "Maschinenzeit", 0.021031, "Log-abgeleitet", "gemessen", "Q-005", "BENCH-C278F4CF; timings.csv; 1,261857 s", "38 Pakete"],
        ["Automatisiert", "A-P03", "Fachliche Abnahme", "Prüfzeit", None, "offen", "offen", "Q-004", "Noch nicht gemessen.", "Benchmark"],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 10, 10)
    for row in range(5, 11):
        if ws.cell(row, 7).value == "offen":
            ws.cell(row, 5).fill = OPEN_FILL
        elif ws.cell(row, 6).value == "geschätzt":
            ws.cell(row, 5).fill = INPUT_FILL
    widths(ws, [16, 12, 34, 20, 14, 16, 14, 18, 52, 20])


def preprocess_summary(workbook) -> None:
    ws = reset_sheet(workbook, "02_PreProcess_Übersicht")
    title(ws, "PreProcess – gruppierte Übersicht", "Für die Ergebnisdarstellung: Eingaben werden gebündelt; die Detailansicht bleibt der Nachweis.", 8)
    header(ws, 4, ["Prozessschritt", "Manuell aktiv (min)", "Manuell Maschine (min)", "Automatisiert aktiv (min)", "Automatisiert Maschine (min)", "Herkunft", "Status", "Quelle / Register-ID", "Hinweis"])
    rows = [
        ["Eingabemodule", "='01_PreProcess_Detail'!E5", "", "", "='01_PreProcess_Detail'!E6+'01_PreProcess_Detail'!E7", "geschätzt / Log", "vorläufig", "Q-003; Q-005", "Manuell: Gesamtaufwand einer neuen Variante."],
        ["Parameter und Varianten", "", "", "", "", "offen", "offen", "Q-004", "Nicht doppelt zum Gesamtaufwand zählen."],
        ["Prüfung / Abnahme", "", "", "", "", "offen", "offen", "Q-004", "Für beide Prozessmodi noch messen."],
        ["Summe bekannte Werte", "=SUM(B5:B7)", "=SUM(C5:C7)", "=SUM(D5:D7)", "=SUM(E5:E7)", "berechnet", "unvollständig", "Q-003; Q-005", "Keine Vergleichskennzahl ableiten."],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 8, 9)
    for cell in ("B5", "E5"):
        ws[cell].fill = INPUT_FILL
    widths(ws, [28, 20, 20, 22, 24, 17, 16, 20, 48])


def core_values(workbook) -> None:
    ws = reset_sheet(workbook, "04_Kernprozess_Einzelwerte")
    title(ws, "Kernprozess – Einzelwerte", "Reine PC-Simulationszeiten aus den bereitgestellten IDA-ICE-Berichten; als manuelle Eingabe dokumentiert.", 10)
    header(ws, 4, ["Variante", "Zonen", "Bereich", "Dauer (s)", "Dauer (min)", "Zeitart", "Herkunft", "Status", "Quelle / Register-ID", "Datum", "Hinweis"])
    rows = [
        ["5Z-Dimensionierung", 5, "Heizlast", 24, "=D5/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-001", "13.08.2026", "IDA-ICE-Bericht"],
        ["5Z-Dimensionierung", 5, "Kühllast", 24, "=D6/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-001", "13.08.2026", "IDA-ICE-Bericht"],
        ["5Z-Dimensionierung", 5, "Energie (ganzjährig)", 294, "=D7/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-001", "13.08.2026", "IDA-ICE-Bericht"],
        ["5Z-Dimensionierung", 5, "Überhitzung", 24, "=D8/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-001", "13.08.2026", "IDA-ICE-Bericht"],
        ["5Z-Dimensionierung", 5, "Summe", "=SUM(D5:D8)", "=SUM(E5:E8)", "Maschinenzeit", "berechnet", "gemessen", "Q-001", "", "6 min 06 s"],
        ["29Z-Dimensionierung", 29, "Heizlast", 109, "=D10/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-002", "12.08.2026", "IDA-ICE-Bericht"],
        ["29Z-Dimensionierung", 29, "Kühllast", 117, "=D11/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-002", "12.08.2026", "IDA-ICE-Bericht"],
        ["29Z-Dimensionierung", 29, "Energie (ganzjährig)", 701, "=D12/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-002", "12.08.2026", "IDA-ICE-Bericht"],
        ["29Z-Dimensionierung", 29, "Überhitzung", 118, "=D13/60", "Maschinenzeit", "manuelle Eingabe", "gemessen", "Q-002", "12.08.2026", "IDA-ICE-Bericht"],
        ["29Z-Dimensionierung", 29, "Summe", "=SUM(D10:D13)", "=SUM(E10:E13)", "Maschinenzeit", "berechnet", "gemessen", "Q-002", "", "17 min 25 s"],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 14, 11)
    for row in (9, 14):
        for cell in ws[row]:
            cell.fill = SECTION_FILL
            cell.font = Font(bold=True)
    widths(ws, [25, 10, 22, 13, 14, 18, 20, 14, 20, 14, 26])


def core_summary(workbook) -> None:
    ws = reset_sheet(workbook, "03_Kernprozess_Übersicht")
    title(ws, "Kernprozess – Übersicht", "Zielprozess: Export/Übergabe → Simulation → Import/Standardisierung. Korrekturen nur bei tatsächlichem Bedarf ergänzen.", 8)
    header(ws, 4, ["Schritt", "Manuell aktiv (min)", "Manuell Maschine (min)", "Automatisiert aktiv (min)", "Automatisiert Maschine (min)", "Herkunft", "Status", "Quelle / Register-ID", "Hinweis"])
    rows = [
        ["Export / Übergabe", "", "", "", "", "offen", "offen", "Q-004", "Für beide Wege noch messen."],
        ["Simulation – 5Z-Dimensionierung", "", "='04_Kernprozess_Einzelwerte'!E9", "", "", "manuelle Eingabe", "gemessen", "Q-001", "Reine PC-Zeit; nicht mit 29Z mitteln."],
        ["Simulation – 29Z-Dimensionierung", "", "='04_Kernprozess_Einzelwerte'!E14", "", "", "manuelle Eingabe", "gemessen", "Q-002", "Reine PC-Zeit; exemplarischer Fallwert."],
        ["Import / Standardisierung", "", "", "", "", "offen", "offen", "Q-004", "Für beide Wege noch messen."],
        ["Prüfung / Korrektur (falls nötig)", "", "", "", "", "offen", "offen", "Q-004", "Nur bei tatsächlichem Anfall erfassen."],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 9, 9)
    widths(ws, [34, 20, 22, 22, 25, 20, 15, 20, 48])


def postprocess(workbook) -> None:
    ws = reset_sheet(workbook, "05_PostProcess_Vorlage")
    title(ws, "PostProcess – Messvorlage", "Die finale Aufteilung wird erst nach weiterer fachlicher Planung festgelegt. Diese Zeilen sind bewusst nur ein anpassbares Erfassungsraster.", 9)
    header(ws, 4, ["Ebene", "Vorgeschlagener Schritt", "Manuell aktiv (min)", "Manuell Maschine (min)", "Automatisiert aktiv (min)", "Automatisiert Maschine (min)", "Herkunft", "Status", "Quelle / Register-ID", "Klärungsnotiz"])
    rows = [
        ["Detail", "Modul / Arbeitsschritt 1", "", "", "", "", "offen", "Messkonzept offen", "Q-004", "Nach gemeinsamer Festlegung umbenennen oder ergänzen."],
        ["Detail", "Modul / Arbeitsschritt 2", "", "", "", "", "offen", "Messkonzept offen", "Q-004", "Keine fachliche Gruppierung vorwegnehmen."],
        ["Detail", "Modul / Arbeitsschritt 3", "", "", "", "", "offen", "Messkonzept offen", "Q-004", "Keine fachliche Gruppierung vorwegnehmen."],
        ["Übersicht", "spätere Gruppierung", "", "", "", "", "offen", "Messkonzept offen", "Q-004", "Erst nach Entscheidung aus Detailzeilen bilden."],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 8, 10)
    for row in range(5, 9):
        for column in range(3, 7):
            ws.cell(row, column).fill = OPEN_FILL
    widths(ws, [14, 34, 20, 22, 22, 25, 16, 22, 20, 50])


def overall(workbook) -> None:
    ws = reset_sheet(workbook, "00_Gesamtübersicht")
    title(ws, "Prozesszeitvergleich – Gesamtübersicht", "Aktueller Messstand: noch nicht vergleichbar. Zeiten und Kosten werden gezeigt, Einsparungen erst nach vollständiger Vergleichbarkeit bewertet.", 11)
    ws["A4"] = "Vergleichbarkeit"
    ws["B4"] = "Nein – Messlücken in beiden Prozesswegen"
    ws["A4"].font = ws["B4"].font = Font(bold=True)
    ws["B4"].fill = OPEN_FILL
    header(ws, 6, ["Prozessphase", "Manuell aktiv (min)", "Manuell Maschine (min)", "Automatisiert aktiv (min)", "Automatisiert Maschine (min)", "Manuell Kosten (€)", "Automatisiert Kosten (€)", "Zeitdifferenz (min)", "Kostendifferenz (€)", "Status", "Quelle / Register-ID", "Hinweis"])
    rows = [
        ["PreProcess", "='02_PreProcess_Übersicht'!B8", "='02_PreProcess_Übersicht'!C8", "='02_PreProcess_Übersicht'!D8", "='02_PreProcess_Übersicht'!E8", "='06_Kosten'!H11", "='06_Kosten'!K11", "", "", "unvollständig", "Q-003; Q-005; Q-006", "Bekannte Werte; noch keine belastbare Gegenüberstellung."],
        ["Kernprozess", "", "", "", "", "='06_Kosten'!H12", "='06_Kosten'!K12", "", "", "variantenbezogen", "Q-001; Q-002; Q-006", "5Z und 29Z sind exemplarisch, nicht zu mitteln."],
        ["PostProcess", "", "", "", "", "='06_Kosten'!H13", "='06_Kosten'!K13", "", "", "Messkonzept offen", "Q-004; Q-006", "Wird erst nach fachlicher Strukturierung gefüllt."],
        ["Gesamt", "=SUM(B7:B9)", "=SUM(C7:C9)", "=SUM(D7:D9)", "=SUM(E7:E9)", "=SUM(F7:F9)", "=SUM(G7:G9)", "=IF(B10+C10+D10+E10=0,\"\",(B10+C10)-(D10+E10))", "=IF(OR(F10=\"\",G10=\"\"),\"\",F10-G10)", "noch nicht vergleichbar", "Q-006", "Differenzen sind nur nach Vergleichbarkeitsfreigabe zu interpretieren."],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 7, 10, 12)
    for column in range(1, 13):
        ws.cell(10, column).fill = SECTION_FILL
        ws.cell(10, column).font = Font(bold=True)
    widths(ws, [20, 20, 22, 22, 25, 19, 22, 20, 20, 22, 24, 56])


def costs(workbook) -> None:
    ws = reset_sheet(workbook, "06_Kosten")
    title(ws, "Kosten – Vorlage und Rechentest", "Gelbe Werte sind ausdrücklich Testannahmen und müssen vor einer Masterarbeitsauswertung durch belegte Werte ersetzt werden.", 11)
    header(ws, 4, ["Parameter", "Wert", "Einheit", "Herkunft", "Status", "Quelle / Register-ID", "Erläuterung", "Manuell", "Automatisiert"])
    parameters = [
        ["Stundensatz", 60, "€/h", "Testannahme", "nicht für Auswertung", "Q-007", "Nur zur Prüfung der Formeln."],
        ["Strompreis", 0.35, "€/kWh", "Testannahme", "nicht für Auswertung", "Q-007", "Vor späterem Kostenvergleich belegen."],
        ["PC-Leistung", 0.25, "kW", "Testannahme", "nicht für Auswertung", "Q-007", "Nur für Maschinenkosten-Test."],
    ]
    for row in parameters:
        ws.append(row)
    for row in range(5, 8):
        ws.cell(row, 2).fill = INPUT_FILL
    header(ws, 10, ["Prozessphase", "Manuell aktiv (min)", "Manuell Maschine (min)", "Auto aktiv (min)", "Auto Maschine (min)", "Manuell Personal (€)", "Manuell Maschine (€)", "Manuell gesamt (€)", "Auto Personal (€)", "Auto Maschine (€)", "Auto gesamt (€)"])
    rows = [
        ["PreProcess", "='00_Gesamtübersicht'!B7", "='00_Gesamtübersicht'!C7", "='00_Gesamtübersicht'!D7", "='00_Gesamtübersicht'!E7", "=IF(B11=\"\",\"\",B11/60*$B$5)", "=IF(C11=\"\",\"\",C11/60*$B$6*$B$7)", "=SUM(F11:G11)", "=IF(D11=\"\",\"\",D11/60*$B$5)", "=IF(E11=\"\",\"\",E11/60*$B$6*$B$7)", "=SUM(I11:J11)"],
        ["Kernprozess", "", "", "", "", "=IF(B12=\"\",\"\",B12/60*$B$5)", "=IF(C12=\"\",\"\",C12/60*$B$6*$B$7)", "=SUM(F12:G12)", "=IF(D12=\"\",\"\",D12/60*$B$5)", "=IF(E12=\"\",\"\",E12/60*$B$6*$B$7)", "=SUM(I12:J12)"],
        ["PostProcess", "", "", "", "", "=IF(B13=\"\",\"\",B13/60*$B$5)", "=IF(C13=\"\",\"\",C13/60*$B$6*$B$7)", "=SUM(F13:G13)", "=IF(D13=\"\",\"\",D13/60*$B$5)", "=IF(E13=\"\",\"\",E13/60*$B$6*$B$7)", "=SUM(I13:J13)"],
        ["Gesamt", "=SUM(B11:B13)", "=SUM(C11:C13)", "=SUM(D11:D13)", "=SUM(E11:E13)", "=SUM(F11:F13)", "=SUM(G11:G13)", "=SUM(H11:H13)", "=SUM(I11:I13)", "=SUM(J11:J13)", "=SUM(K11:K13)"],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 7, 9)
    table_style(ws, 11, 14, 11)
    for cell in ws[14]:
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True)
    widths(ws, [22, 18, 18, 18, 22, 20, 42, 18, 20, 20, 20])


def assumptions(workbook) -> None:
    ws = reset_sheet(workbook, "08_Messannahmen")
    title(ws, "Messannahmen und Nachweise", "Diese Dokumentation macht deutlich, was gemessen, manuell eingetragen, geschätzt oder noch offen ist.", 6)
    header(ws, 4, ["Thema", "Festlegung / Messwert", "Herkunft", "Status", "Quelle / Register-ID", "Auswirkung", "Nachweis / nächste Aktion"])
    rows = [
        ["Prozessgrenzen", "PreProcess, Kernprozess und PostProcess getrennt", "P030", "festgelegt", "Q-006", "Keine Vermischung aktiver und Maschinenzeit", "P030 Prozessauswertung"],
        ["Neue Variante", "1–4 h aktive Eingabezeit insgesamt", "Nutzerangabe", "geschätzt", "Q-003", "Nicht auf mehrere Module doppelt anrechnen", "Später mit Ist-Zeit ersetzen"],
        ["Folgevariante", "Reduziert durch Kopie; nicht separat gemessen", "Nutzerangabe", "offen", "Q-004", "Keine geschätzte Kostenrechnung", "Bei nächstem Ablauf erfassen"],
        ["5Z-Simulation", "366 s / 6 min 06 s", "IDA-ICE-Bericht", "gemessen", "Q-001", "Exemplarischer PC-Zeitwert", "13.08.2026"],
        ["29Z-Simulation", "1.045 s / 17 min 25 s", "IDA-ICE-Bericht", "gemessen", "Q-002", "Exemplarischer PC-Zeitwert", "12.08.2026"],
        ["PostProcess", "Modulstruktur noch nicht festgelegt", "Nutzerentscheidung", "offen", "Q-004", "Keine Gruppierung vorwegnehmen", "Später gemeinsam festlegen"],
        ["Vergleichbarkeit", "Noch nicht gegeben", "Messstand", "offen", "Q-006", "Keine belastbare Einsparung ausweisen", "Gleiche Grenzen und Artefakte messen"],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 11, 7)
    widths(ws, [22, 36, 20, 16, 20, 42, 40])


def sources_register(workbook) -> None:
    ws = reset_sheet(workbook, "09_Quellenregister")
    title(ws, "Quellenregister und Begründungen", "Alle Register-IDs aus den übrigen Blättern verweisen hierher. Sie dokumentieren Herkunft, Begründung, Gültigkeitsbereich und erforderliche Aktualisierung.", 9)
    header(ws, 4, ["Register-ID", "Betroffene Werte / Annahmen", "Quellentyp", "Quelle / Fundstelle", "Begründung der Verwendung", "Gültigkeitsbereich", "Status", "Grenze / Risiko", "Nächste Aktion"])
    rows = [
        ["Q-001", "5Z-Dimensionierung: Heizlast, Kühllast, Energie, Überhitzung", "IDA-ICE-Bericht / manuelle Eingabe", "Bereitgestellter Screenshot; Simulationen vom 13.08.2026", "Direkt abgelesene PC-Simulationsdauern aus den vier Berichten.", "5-Zonen-Fall, nur Kernprozess-Simulation", "gemessen", "Keine aktive Nutzer- oder Wartezeit enthalten.", "Bei Verfügbarkeit Bericht/Log als Nachweis ablegen."],
        ["Q-002", "29Z-Dimensionierung: Heizlast, Kühllast, Energie, Überhitzung", "IDA-ICE-Bericht / manuelle Eingabe", "Bereitgestellter Screenshot; Simulationen vom 12.08.2026", "Direkt abgelesene PC-Simulationsdauern aus den vier Berichten.", "29-Zonen-Fall, nur Kernprozess-Simulation", "gemessen", "Kein allgemeiner Leistungsnachweis; nicht mit 5Z mitteln.", "Bei Verfügbarkeit Bericht/Log als Nachweis ablegen."],
        ["Q-003", "1–4 h für vollständige Eingabe einer neuen Variante", "Nutzerangabe / Schätzung", "Chat-Angabe des Nutzers", "Einziger derzeit vorliegender Erfahrungswert für den aktiven Gesamtaufwand.", "Nur neue Variante; keine mehrfachen Teilmodulwerte", "geschätzt", "Spanne, kein beobachteter Zeitstempel; 150 min ist nur Rechentest.", "Bei nächstem vollständigen Ablauf mit Start/Ende messen."],
        ["Q-004", "Offene Pre-, Kern- und PostProcess-Schritte; Folgevarianten", "Offener Messpunkt / Nutzerentscheidung", "Aktueller Messstand und Gesprächsfestlegungen", "Verhindert unbelegte Schätzungen und hält Messlücken sichtbar.", "Alle noch nicht separat erfassten Prozessschritte", "offen", "Nicht für Einsparungs- oder Kosteninterpretationen verwenden.", "Messkonzept bzw. Ist-Zeit vor Vergleich ergänzen."],
        ["Q-005", "Automatisierter Workspace- und PreProcess-Benchmark", "Benchmark-Log", "BENCH-C278F4CF; timings.csv", "Technisch erfasste Laufzeiten der vorhandenen PreProcess-Ausführung.", "Benchmark mit 38 Paketen; Maschinenzeit", "Log-abgeleitet", "Deckt keine Eingabe oder fachliche Abnahme ab.", "Bei neuem Benchmark Log und Konfiguration referenzieren."],
        ["Q-006", "Prozessgrenzen und Vergleichbarkeitsregel", "Projektplan", "P030 research_tools Prozessmessung und Vergleichsauswertung", "Kanonische Methodik: getrennte Zeitarten und gleiche Prozessgrenzen.", "Gesamte Prozessauswertung", "festgelegt", "Vergleichskennzahlen erst bei vergleichbarer Messbasis.", "Vor Ergebnisinterpretation Vergleichbarkeit prüfen."],
        ["Q-007", "Stundensatz, Strompreis und PC-Leistung", "Testannahme", "Arbeitsmappe; keine externe Quelle", "Nur zum Prüfen von Formeln und Tabellenverknüpfungen eingesetzt.", "Kostenblatt, ausschließlich Rechentest", "Testwert", "Nicht als Masterarbeitswert oder Kostenersparnis verwenden.", "Durch belegte Annahmen nach OP-009 ersetzen."],
    ]
    for row in rows:
        ws.append(row)
    table_style(ws, 5, 11, 9)
    widths(ws, [16, 42, 28, 42, 44, 36, 16, 44, 42])


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    # The historical sheets Manuell, Prozess automatisiert, Kosten and Vergleich remain untouched.
    overall(workbook)
    preprocess_detail(workbook)
    preprocess_summary(workbook)
    core_summary(workbook)
    core_values(workbook)
    postprocess(workbook)
    costs(workbook)
    assumptions(workbook)
    sources_register(workbook)

    for name in SHEETS:
        workbook[name].sheet_view.showGridLines = False

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(WORKBOOK_PATH)
    print(f"Aktualisiert: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
