"""Create ignored local building catalogs from the three approved workbook sections.

The script deliberately reads no other workbook sections. Its outputs remain
ignored under ``config/ma_database/catalogs`` and must not be published.
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import load_workbook

SOURCE_PATH = Path("data/project_inbox/new/unknown/Gebäude-Zonenansicht.xlsx")
OUTPUT_DIRECTORY = Path("config/ma_database/catalogs/v0.1.0")
SHEET_NAME = "DIM - Räume"


def _value(value):
    return value.strip() if isinstance(value, str) else value


def _layer_pairs(values: list[object], start_index: int) -> list[dict[str, object]]:
    layers = []
    for index in range(start_index, len(values), 2):
        material = _value(values[index])
        thickness = values[index + 1] if index + 1 < len(values) else None
        if material is not None or thickness is not None:
            layers.append({"layer_no": len(layers) + 1, "material_name": material or "", "thickness_m": thickness})
    return layers


def _nonempty_rows(worksheet, start: int, end: int) -> list[list[object]]:
    return [
        [_value(cell.value) for cell in row]
        for row in worksheet.iter_rows(min_row=start, max_row=end, values_only=False)
        if _value(row[0].value) not in (None, "")
    ]


def _write_catalog(filename: str, catalog_type: str, records: list[dict[str, object]]) -> None:
    OUTPUT_DIRECTORY.mkdir(parents=True, exist_ok=True)
    payload = {"schema_version": "1.0", "catalog_type": catalog_type, "records": records}
    (OUTPUT_DIRECTORY / filename).write_text(
        yaml.safe_dump(payload, allow_unicode=True, sort_keys=False, width=120), encoding="utf-8"
    )


def main() -> None:
    workbook = load_workbook(SOURCE_PATH, data_only=True, read_only=True)
    worksheet = workbook[SHEET_NAME]

    surface_rows = _nonempty_rows(worksheet, 74, 104)
    surfaces = [
        {
            "name": row[0],
            "surface_id": f"BSURF-{source_row:04d}",
            "group": row[1],
            "surface_type": row[2],
            "wetted_area_m2": row[3],
            "connected_to": row[4],
            "azimuth_deg": row[5],
            "slope_deg": row[6],
            "construction_name": row[7],
            "u_value_w_m2k": row[8],
            "thickness_m": row[9],
            "layers": _layer_pairs(row, 10),
        }
        for source_row, row in zip(range(74, 105), surface_rows, strict=True)
    ]

    construction_rows = _nonempty_rows(worksheet, 127, 138)
    constructions = [
        {
            "name": row[0],
            "wall_construction_id": f"BWCON-{source_row:04d}",
            "u_value_w_m2k": row[1],
            "thickness_m": row[2],
            "layers": _layer_pairs(row, 3),
        }
        for source_row, row in zip(range(127, 139), construction_rows, strict=True)
    ]

    material_rows = _nonempty_rows(worksheet, 142, 172)
    materials = [
        {
            "name": row[0],
            "material_id": f"BMAT-{source_row:04d}",
            "heat_conductivity_w_mk": row[1],
            "density_kg_m3": row[2],
            "specific_heat_j_kgk": row[3],
            "total_area_m2": row[4],
            "total_volume_m3": row[5],
            "total_mass_kg": row[6],
        }
        for source_row, row in zip(range(142, 173), material_rows, strict=True)
    ]

    _write_catalog("building_materials.yaml", "building_materials", materials)
    _write_catalog("building_wall_constructions.yaml", "building_wall_constructions", constructions)
    _write_catalog("building_surfaces.yaml", "building_surfaces", surfaces)
    print(f"Wrote {len(materials)} materials, {len(constructions)} wall constructions, {len(surfaces)} surfaces.")


if __name__ == "__main__":
    main()
