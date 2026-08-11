import hashlib
import json
from copy import deepcopy
from dataclasses import replace

import pandas as pd
import pytest
import yaml
from openpyxl import Workbook

from ma_building import load_building_excel_catalog
from ma_parameters import (
    DIN_USAGE_PROFILE_METADATA,
    build_small_office_5z_v1_baseline_parameter_snapshot,
    reference_dimensioning_parameter_fingerprint,
    suggest_usage_profile_id,
    variation_specification_source_fingerprint,
)
from ma_simulation_setup import materialize_project_setup_packages
from ma_technical import (
    load_business_integration_lod1_technical_spec,
    validate_technical_spec,
)
from ma_ui.streamlit_app.module_views.building_view import (
    building_excel_selection_payload,
)
from ma_ui.streamlit_app.module_views.dimensioning_view import (
    manual_reference_load_rows,
    validate_manual_reference_load_rows,
)
from ma_ui.streamlit_app.module_views.parameters_view import validate_parameter_project_payload
from ma_ui.streamlit_app.module_views.simulation_setup_view import simulation_setup_output_root
from ma_ui.streamlit_app.pages.variants import _dimensioning_complete, naming_preview_rows_for_context
from ma_ui.streamlit_app.state import (
    clear_workspace_draft,
    mark_workspace_draft,
    open_workspace_drafts,
    small_office_v1_uses_reference_zone_model,
)
from ma_variants import (
    build_small_office_candidate_rows,
    candidate_simulation_setup,
    candidate_source_is_current,
    load_small_office_v1_study,
    materialize_zonal_capacities,
    select_candidate_ids,
    small_office_study_case_rows,
    source_fingerprint,
    variation_specification_is_current,
    verify_candidate_rows,
)
from ma_zones import (
    build_small_office_29z_draft,
    load_business_integration_lod1_zone_spec,
    load_small_office_5z_endvariant_02_zone_spec,
    zone_specification_to_dict,
)


def _study_variation_specification(study) -> dict[str, object]:
    return {
        "status": "current",
        "study_contract": {
            "study_id": study.study_id,
            "enabled_dimensions": [
                "temperature_setpoint_bands",
                "coupled_heating_cooling_capacity_factors",
                "weather_ofat",
                "occupancy_ofat",
            ],
        },
    }


def test_building_excel_catalog_reads_only_overview_records(tmp_path):
    source = tmp_path / "components.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Übersicht"
    sheet.append(["Hinweis"])
    sheet.append([])
    sheet.append([])
    sheet.append(["Bauteil-ID", "Bauteil", "Prüfstatus"])
    sheet.append(["AW-001", "Außenwand", "validated"])
    sheet.append(["TEXT", "kein Datensatz", ""])
    workbook.save(source)

    catalog = load_building_excel_catalog("components", source)

    assert catalog.source_path == source.resolve()
    assert len(catalog.source_sha256) == 64
    assert catalog.rows == ({"Bauteil-ID": "AW-001", "Bauteil": "Außenwand", "Prüfstatus": "validated"},)
    project_copy = building_excel_selection_payload(
        catalog,
        catalog.rows[0],
        target_id="ELEMENT-001",
        target_group={"element_type": "wall", "construction_code": "AW"},
        scope="element",
    )
    assert project_copy["catalog_record"] == catalog.rows[0]
    assert project_copy["source_sha256"] == catalog.source_sha256


def test_building_excel_catalog_rejects_duplicate_record_ids(tmp_path):
    source = tmp_path / "components.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Übersicht"
    sheet.append(["Hinweis"])
    sheet.append([])
    sheet.append([])
    sheet.append(["Bauteil-ID", "Bauteil"])
    sheet.append(["AW-001", "Außenwand A"])
    sheet.append(["AW-001", "Außenwand B"])
    workbook.save(source)

    with pytest.raises(ValueError, match="doppelte Datensatz-IDs"):
        load_building_excel_catalog("components", source)


def test_building_excel_catalog_rejects_duplicate_headers(tmp_path):
    source = tmp_path / "components.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Übersicht"
    sheet.append(["Hinweis"])
    sheet.append([])
    sheet.append([])
    sheet.append(["Bauteil-ID", "Bauteil", "Bauteil"])
    sheet.append(["AW-001", "Außenwand", "Duplikat"])
    workbook.save(source)

    with pytest.raises(ValueError, match="doppelte Tabellenkoepfe"):
        load_building_excel_catalog("components", source)


def test_29z_draft_has_one_zone_per_ifc_room_without_5z_values():
    draft = build_small_office_29z_draft()

    assert len(draft.zones) == 29
    assert all(len(zone.source_space_ids) == 1 for zone in draft.zones)
    assert len({zone.source_space_ids[0] for zone in draft.zones}) == 29
    assert draft.zones[0].name == "Space 001"
    assert draft.zones[0].zone_id == "ZONE-29Z-001"
    assert all(zone.usage_profile_id == "" for zone in draft.zones)
    assert all(zone.heating_setpoint_c == 0 for zone in draft.zones)
    assert all(zone.cooling_setpoint_c == 0 for zone in draft.zones)
    assert len(DIN_USAGE_PROFILE_METADATA) == 43
    assert DIN_USAGE_PROFILE_METADATA[0].edition == "2025-10"
    assert suggest_usage_profile_id("Office") == "DIN18599-A01"
    assert suggest_usage_profile_id("unbekannte Nutzung") is None


def test_lite_technical_validation_rejects_unknown_schema_version():
    spec = replace(
        load_business_integration_lod1_technical_spec(),
        schema_version="9.9",
    )

    result = validate_technical_spec(spec)

    assert any(message.code == "TECHNICAL_SCHEMA_VERSION_INVALID" for message in result.messages)


def test_project_studies_create_both_directions_and_three_cases():
    study = load_small_office_v1_study()
    study_cases = small_office_study_case_rows(study)
    candidates = build_small_office_candidate_rows(
        study,
        _study_variation_specification(study),
    )
    verified = verify_candidate_rows(candidates, reference_dimensioning_complete=True)

    assert {row["study_direction"] for row in study_cases} == {"optimization", "sensitivity"}
    assert len(study_cases) == 3
    assert len(candidates) == 38
    assert sum(row["study_direction"] == "optimization" for row in candidates) == 30
    assert all(row["status"] == "valid" for row in verified)
    assert {row["study_direction_id"] for row in study_cases} == {
        "SD-OPTIMIZATION",
        "SD-SENSITIVITY",
    }
    assert {row["values"]["parent_variant_id"] for row in candidates if row["study_direction"] == "sensitivity"} == {
        "OPT-SB01-F100"
    }
    valid_ids = tuple(str(row["candidate_id"]) for row in verified)
    assert select_candidate_ids(valid_ids, mode="zufaellig", count=3, seed=42) == select_candidate_ids(
        valid_ids,
        mode="zufaellig",
        count=3,
        seed=42,
    )


def test_random_156_study_creates_156_candidates_and_reproducible_selection():
    study = load_small_office_v1_study("config/ma_variants/studies/small_office_v1_random_156.yaml")
    candidates = build_small_office_candidate_rows(study, _study_variation_specification(study))
    optimization_ids = tuple(
        str(row["candidate_id"])
        for row in candidates
        if row["study_direction"] == "optimization"
    )
    assert len(optimization_ids) == 156
    assert select_candidate_ids(optimization_ids, mode="zufaellig", count=50, seed=20260806) == select_candidate_ids(
        optimization_ids,
        mode="zufaellig",
        count=50,
        seed=20260806,
    )


def test_simulation_setup_output_roots_keep_tests_and_projects_separate():
    test_root = simulation_setup_output_root("PRJ-TEST-001", test_only=True)
    project_root = simulation_setup_output_root("PRJ-REAL-001", test_only=False)

    assert test_root.as_posix().endswith("data/test_output/PRJ-TEST-001/simulation_setup")
    assert project_root.as_posix().endswith("data/project_output/PRJ-REAL-001/simulation_setup")

@pytest.mark.parametrize(
    "dimensions",
    [
        [
            "temperature_setpoint_bands",
            "coupled_heating_cooling_capacity_factors",
            "weather_ofat",
            "occupancy_ofat",
            "unknown_fifth_dimension",
        ],
        [
            "temperature_setpoint_bands",
            "coupled_heating_cooling_capacity_factors",
            "weather_ofat",
            "weather_ofat",
        ],
    ],
)
def test_project_study_rejects_unknown_or_duplicate_dimensions(dimensions):
    study = load_small_office_v1_study()
    specification = _study_variation_specification(study)
    specification["study_contract"]["enabled_dimensions"] = dimensions

    with pytest.raises(ValueError, match="exakt und eindeutig"):
        build_small_office_candidate_rows(study, specification)


def test_project_study_materializes_zonal_capacity_and_setup():
    study = load_small_office_v1_study()
    candidate = build_small_office_candidate_rows(
        study,
        _study_variation_specification(study),
    )[1]

    capacities = materialize_zonal_capacities(
        candidate,
        [
            {
                "zone_id": "ZONE-001",
                "zone_name": "Zone 1",
                "heating_load_w": 1000.0,
                "cooling_load_w": 800.0,
            }
        ],
    )
    setup = candidate_simulation_setup(study, candidate)

    assert capacities[0]["available_heating_capacity_w"] == 900.0
    assert capacities[0]["available_cooling_capacity_w"] == 720.0
    assert setup["weather_key"] == "TRY_FFM_2015_JAHR"
    assert setup["occupancy_start_hour"] == 7.0
    assert setup["simulation_timestep_seconds"] == 3600
    assert setup["calendar_definition"] == "TRY_non_leap_standard_year_8760"
    assert setup["weather_source_reference"]["study_record_fingerprint"]
    assert setup["weather_source_reference"]["source_status"] in {
        "resolved_local_file",
        "source_resolution_required_before_simulation",
    }


def test_variation_contract_requires_explicit_current_study_approval():
    study = load_small_office_v1_study()
    baseline = build_small_office_5z_v1_baseline_parameter_snapshot()
    study_contract = _study_variation_specification(study)["study_contract"]
    parameter_payload = {
        "rules": [],
        "variation_spans": [],
        "variation_specification": {
            "status": "draft",
            "study_contract": study_contract,
        },
    }
    parameter_payload["variation_specification"]["source_fingerprint"] = variation_specification_source_fingerprint(
        baseline,
        rules=[],
        variation_spans=[],
        study_contract=study_contract,
    )

    assert not variation_specification_is_current(parameter_payload, baseline, study)
    parameter_payload["variation_specification"]["status"] = "current"
    assert variation_specification_is_current(parameter_payload, baseline, study)


def test_stale_candidate_source_cannot_be_relabelled_current():
    payload = {
        "candidates": [{"candidate_id": "OLD"}],
        "source_fingerprint": "old",
    }

    assert not candidate_source_is_current(payload, "new")
    assert candidate_source_is_current(payload, "old")


def test_naming_preview_is_bound_to_exact_context():
    context = {
        "project_id": "PRJ-001",
        "study_case_id": "SC-001",
        "candidate_ids": ["VAR-001"],
        "source_fingerprint": "source",
        "naming_profile_reference": "naming.yaml",
    }
    preview = {"context": context, "rows": [{"candidate_id": "VAR-001"}]}

    assert naming_preview_rows_for_context(preview, context) == preview["rows"]
    assert (
        naming_preview_rows_for_context(
            preview,
            {**context, "project_id": "PRJ-002"},
        )
        is None
    )


def test_parameter_project_payload_requires_scoped_rules_and_reference():
    errors = validate_parameter_project_payload(
        {
            "reference": {"snapshot_id": "PARAM-001"},
            "rules": [
                {
                    "rule_id": "RULE-001",
                    "title": "StudyCase-Regel",
                    "scope_type": "study_case",
                    "scope_id": "",
                }
            ],
            "variation_spans": [],
        }
    )

    assert errors == ("Regel 1 braucht eine StudyDirection- oder StudyCase-ID.",)


def test_parameter_project_payload_rejects_invalid_numeric_span():
    errors = validate_parameter_project_payload(
        {
            "reference": {"snapshot_id": "PARAM-001"},
            "rules": [],
            "variation_spans": [
                {
                    "parameter_key": "zone.heating_setpoint_c",
                    "enabled": True,
                    "value_form": "Min/Max/Schritt",
                    "minimum": 10,
                    "maximum": 1,
                    "step": 0,
                }
            ],
        }
    )

    assert any("Minimum" in error for error in errors)
    assert any("Schritt" in error for error in errors)


def test_manual_reference_loads_reject_reordered_zone_names():
    zone_spec = load_business_integration_lod1_zone_spec()
    rows = manual_reference_load_rows(zone_spec, None)
    rows[0]["Zone"] = "Manipulierte Zone"
    rows[0]["Heizlast [W]"] = 1000.0
    rows[0]["Kuehllast [W]"] = 800.0

    with pytest.raises(ValueError, match="sortiert oder veraendert"):
        validate_manual_reference_load_rows(zone_spec, pd.DataFrame(rows))


def test_dimensioning_gate_rejects_stale_reference_parameter_fingerprint():
    baseline = build_small_office_5z_v1_baseline_parameter_snapshot()
    zone_spec = load_small_office_5z_endvariant_02_zone_spec()
    parameter_payload = {"reference": {"snapshot_id": baseline.snapshot_id}}
    zone_payload = zone_specification_to_dict(zone_spec)
    zone_hash = hashlib.sha256(
        json.dumps(
            zone_payload,
            sort_keys=True,
            ensure_ascii=False,
            default=str,
        ).encode("utf-8")
    ).hexdigest()
    payload = {
        "zone_model_id": zone_spec.zone_model_id,
        "zone_model_hash": zone_hash,
        "reference_parameter_fingerprint": "stale",
        "zone_loads": [
            {
                "zone_id": zone.zone_id,
                "heating_load_w": 1000.0,
                "cooling_load_w": 800.0,
            }
            for zone in zone_spec.zones
        ],
        "ida_source": {
            "ida_version": "5.0",
            "model_id": "MODEL-001",
            "run_id": "RUN-001",
            "source_file_name": "model.idm",
            "source_file_sha256": "a" * 64,
            "heating_load_definition": "zone_heating_load",
            "cooling_load_definition": "sensible_zone_load",
            "maximum_definition": "individual_zone_maximum",
            "design_conditions": "annual",
            "responsible": "Author",
            "review_status": "reviewed",
            "reviewer": "Reviewer",
            "reviewed_at": "2026-07-27T12:00:00+02:00",
            "review_note": "checked",
            "source_classification": "externally_simulated_result",
        },
    }

    assert not _dimensioning_complete(
        payload,
        zone_spec,
        baseline,
        parameter_payload,
    )
    payload["reference_parameter_fingerprint"] = reference_dimensioning_parameter_fingerprint(
        baseline,
        parameter_payload,
    )
    assert _dimensioning_complete(
        payload,
        zone_spec,
        baseline,
        parameter_payload,
    )


def test_open_workspace_drafts_are_deduplicated_and_clearable():
    session_state: dict[str, object] = {}

    mark_workspace_draft(session_state, "ma_building")
    mark_workspace_draft(session_state, "ma_building")
    mark_workspace_draft(session_state, "ma_parameters")

    assert open_workspace_drafts(session_state) == ("ma_building", "ma_parameters")
    clear_workspace_draft(session_state, "ma_building")
    assert open_workspace_drafts(session_state) == ("ma_parameters",)


def test_small_office_v1_downstream_guard_rejects_29z():
    assert small_office_v1_uses_reference_zone_model({"active_model": "5Z"})
    assert not small_office_v1_uses_reference_zone_model({"active_model": "29Z"})


def test_simulation_setup_materializes_only_current_confirmed_packages(
    tmp_path,
    monkeypatch,
):
    weather_file = tmp_path / "weather.dat"
    weather_file.write_text("synthetic weather source", encoding="utf-8")
    weather_hash = hashlib.sha256(weather_file.read_bytes()).hexdigest()
    selection_contract = {
        "study_case_id": "SC-001",
        "mode": "manuell",
        "candidate_ids": ["VAR-001"],
        "random_seed": None,
        "source_fingerprint": "abc",
    }
    selection_fingerprint = source_fingerprint(selection_contract)
    selection_id = f"SEL-SC-001-{selection_fingerprint[:12]}"
    package = {
        "schema_version": "1.0",
        "variant_id": "VAR-001",
        "variant_name": "V_001",
        "study_id": "STUDY-001",
        "study_case_id": "SC-001",
        "study_direction_id": "SD-001",
        "selection_id": selection_id,
        "selection_reference": {
            **selection_contract,
            "selection_id": selection_id,
            "selection_fingerprint": selection_fingerprint,
            "study_direction_id": "SD-001",
        },
        "baseline_reference": {"snapshot_id": "PARAM-001"},
        "parameter_reference": {"project_configuration_fingerprint": "parameter"},
        "zone_model_reference": {
            "zone_model_id": "ZONE-MODEL-001",
            "content_hash": "zone",
            "zone_ids": ["ZONE-001"],
        },
        "dimensioning_reference": {"content_hash": "dimensioning"},
        "capacity_strategy": "fixed_reference_21_24_zonal_capacity",
        "zonal_capacities": [
            {
                "zone_id": "ZONE-001",
                "reference_heating_load_w": 1000.0,
                "reference_cooling_load_w": 800.0,
                "heating_factor": 1.0,
                "cooling_factor": 1.0,
                "available_heating_capacity_w": 1000.0,
                "available_cooling_capacity_w": 800.0,
            }
        ],
        "simulation_setup": {
            "weather_key": "TRY_FFM_2015_JAHR",
            "weather_label": "Frankfurt 2015 Jahr",
            "weather_analysis_supported": True,
            "occupancy_schedule_key": "OCC_REF",
            "occupancy_start_hour": 7.0,
            "occupancy_end_hour": 18.0,
            "same_values_for_all_zones": True,
            "time_zone": "Europe/Berlin",
            "simulation_period": "annual",
            "simulation_start": "2015-01-01T00:00:00",
            "simulation_end": "2015-12-31T23:00:00",
            "calendar_definition": "TRY_non_leap_standard_year_8760",
            "daylight_saving_time": False,
            "simulation_timestep_seconds": 3600,
            "weather_source_reference": {
                "source_path": str(weather_file),
                "source_revision": "test",
                "source_file_sha256": weather_hash,
                "study_record_fingerprint": "b" * 64,
                "source_status": "resolved_local_file",
            },
        },
        "output_requirements": [{"profile_id": "OUT-LOAD"}],
        "source_fingerprint": "abc",
        "status": "confirmed",
    }

    paths = materialize_project_setup_packages(
        output_root=tmp_path,
        run_group_id="RUN-GROUP-001",
        project_id="PRJ-000001",
        simulation_program_key="ida_ice",
        variant_packages=[package],
        source_fingerprint="abc",
        study_label="Teststudie",
        test_only=True,
        technical_timings=[
            {
                "stage": "candidate_generation",
                "status": "success",
                "duration_seconds": 0.125,
                "recorded_at": "2026-08-06T10:00:00+00:00",
                "details": "156 Kandidaten erzeugt",
            }
        ],
    )

    assert len(paths) == 1
    assert (paths[0] / "run_manifest.yaml").is_file()
    assert (paths[0] / "variant_package.yaml").is_file()
    assert (paths[0] / "simulation_setup.yaml").is_file()
    assert (paths[0].parent / "selection_manifest.yaml").is_file()
    assert (paths[0].parent / "run_summary.yaml").is_file()
    assert (paths[0].parent / "timings.yaml").is_file()
    assert (paths[0].parent / "timings.csv").is_file()
    assert not tuple(tmp_path.glob("RUN-GROUP-001-staging-*"))
    setup_payload = yaml.safe_load((paths[0] / "simulation_setup.yaml").read_text(encoding="utf-8"))
    assert setup_payload["run_group_id"] == "RUN-GROUP-001"
    assert setup_payload["selection_id"] == selection_id
    assert setup_payload["weather"]["weather_key"] == "TRY_FFM_2015_JAHR"
    assert setup_payload["zonal_capacities"][0]["available_heating_capacity_w"] == 1000.0
    selection_payload = yaml.safe_load((paths[0].parent / "selection_manifest.yaml").read_text(encoding="utf-8"))
    assert selection_payload["status"] == "prepared_for_manual_simulation"
    summary_payload = yaml.safe_load((paths[0].parent / "run_summary.yaml").read_text(encoding="utf-8"))
    assert summary_payload["study_label"] == "Teststudie"
    assert summary_payload["test_only"] is True
    assert summary_payload["selected_variant_count"] == 1
    timing_payload = yaml.safe_load((paths[0].parent / "timings.yaml").read_text(encoding="utf-8"))
    assert timing_payload["timings"][0]["stage"] == "candidate_generation"
    assert any(row["stage"] == "simulation_setup_materialization" for row in timing_payload["timings"])

    def assert_invalid_package(
        invalid_package: dict[str, object],
        *,
        run_group_suffix: str,
        match: str,
    ) -> None:
        with pytest.raises(ValueError, match=match):
            materialize_project_setup_packages(
                output_root=tmp_path,
                run_group_id=f"RUN-GROUP-{run_group_suffix}",
                project_id="PRJ-000001",
                simulation_program_key="ida_ice",
                variant_packages=[invalid_package],
                source_fingerprint="abc",
            )

    contradictory_package = deepcopy(package)
    contradictory_package["selection_id"] = "SEL-CONTRADICTING"
    assert_invalid_package(
        contradictory_package,
        run_group_suffix="CONTRADICTING",
        match="Selection-ID",
    )

    duplicate_candidates_package = deepcopy(package)
    duplicate_candidates_package["selection_reference"]["candidate_ids"] = [
        "VAR-001",
        "VAR-001",
    ]
    assert_invalid_package(
        duplicate_candidates_package,
        run_group_suffix="DUPLICATE-CANDIDATES",
        match="materialisierte Varianten",
    )

    wrong_case_package = deepcopy(package)
    wrong_case_package["study_case_id"] = "SC-OTHER"
    assert_invalid_package(
        wrong_case_package,
        run_group_suffix="WRONG-CASE",
        match="Selection-StudyCase",
    )

    wrong_direction_package = deepcopy(package)
    wrong_direction_package["study_direction_id"] = "SD-OTHER"
    assert_invalid_package(
        wrong_direction_package,
        run_group_suffix="WRONG-DIRECTION",
        match="Selection-StudyDirection",
    )

    wrong_fingerprint_package = deepcopy(package)
    wrong_fingerprint_package["selection_reference"]["selection_fingerprint"] = "invalid"
    assert_invalid_package(
        wrong_fingerprint_package,
        run_group_suffix="WRONG-FINGERPRINT",
        match="Selection-Fingerprint",
    )

    wrong_derived_id_package = deepcopy(package)
    wrong_derived_id_package["selection_id"] = "SEL-SC-001-WRONG"
    wrong_derived_id_package["selection_reference"]["selection_id"] = "SEL-SC-001-WRONG"
    assert_invalid_package(
        wrong_derived_id_package,
        run_group_suffix="WRONG-DERIVED-ID",
        match="Selection-ID ist inkonsistent",
    )

    non_string_id_package = deepcopy(package)
    non_string_id_package["study_direction_id"] = 1
    non_string_id_package["selection_reference"]["study_direction_id"] = "1"
    assert_invalid_package(
        non_string_id_package,
        run_group_suffix="NONSTRING-ID",
        match="nichtleere Strings",
    )

    wrong_calendar_package = deepcopy(package)
    wrong_calendar_package["simulation_setup"]["time_zone"] = "UTC"
    assert_invalid_package(
        wrong_calendar_package,
        run_group_suffix="WRONG-TIME-ZONE",
        match="Europe/Berlin",
    )

    wrong_weather_hash_package = deepcopy(package)
    wrong_weather_hash_package["simulation_setup"]["weather_source_reference"]["source_file_sha256"] = "c" * 64
    assert_invalid_package(
        wrong_weather_hash_package,
        run_group_suffix="WRONG-WEATHER-HASH",
        match="gespeicherten SHA-256",
    )

    incomplete_weather_package = deepcopy(package)
    incomplete_weather_package["simulation_setup"]["weather_source_reference"].update(
        {
            "source_path": "",
            "source_file_sha256": "",
            "source_status": "source_resolution_required_before_simulation",
        }
    )
    incomplete_paths = materialize_project_setup_packages(
        output_root=tmp_path,
        run_group_id="RUN-GROUP-WEATHER-INCOMPLETE",
        project_id="PRJ-000001",
        simulation_program_key="ida_ice",
        variant_packages=[incomplete_weather_package],
        source_fingerprint="abc",
    )
    incomplete_manifest = yaml.safe_load((incomplete_paths[0] / "run_manifest.yaml").read_text(encoding="utf-8"))
    incomplete_setup = yaml.safe_load((incomplete_paths[0] / "simulation_setup.yaml").read_text(encoding="utf-8"))
    assert incomplete_manifest["status"] == "preparation_incomplete_weather_source"
    assert incomplete_setup["next_action"] == "resolve_weather_source_before_manual_simulation"

    with pytest.raises(ValueError, match="aktualisierungsbeduerftig"):
        materialize_project_setup_packages(
            output_root=tmp_path,
            run_group_id="RUN-GROUP-002",
            project_id="PRJ-000001",
            simulation_program_key="ida_ice",
            variant_packages=[package],
            source_fingerprint="changed",
        )

    import ma_simulation_setup.project_packages as project_packages

    original_writer = project_packages._write_yaml_new

    def fail_during_variant_write(path, payload):
        if path.name == "variant_package.yaml":
            raise OSError("synthetic write failure")
        original_writer(path, payload)

    monkeypatch.setattr(
        project_packages,
        "_write_yaml_new",
        fail_during_variant_write,
    )
    with pytest.raises(OSError, match="synthetic write failure"):
        materialize_project_setup_packages(
            output_root=tmp_path,
            run_group_id="RUN-GROUP-ROLLBACK",
            project_id="PRJ-000001",
            simulation_program_key="ida_ice",
            variant_packages=[package],
            source_fingerprint="abc",
        )
    assert not (tmp_path / "RUN-GROUP-ROLLBACK").exists()
    assert not tuple(tmp_path.glob("RUN-GROUP-ROLLBACK-staging-*"))
